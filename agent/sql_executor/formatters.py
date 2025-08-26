"""
Result formatters for SQL Executor
Handles multiple output formats: JSON, CSV, HTML, Excel
"""

import json
import csv
import io
from typing import List, Dict, Any, Optional, Union
from datetime import datetime, date
import pandas as pd
from abc import ABC, abstractmethod

from .models import ExecutionResult, ColumnInfo, OutputFormat
from .config import get_executor_config

class BaseFormatter(ABC):
    """Abstract base formatter for all output formats"""
    
    def __init__(self):
        self.config = get_executor_config()
    
    @abstractmethod
    def format(self, result: ExecutionResult) -> Union[str, bytes]:
        """Format execution result to specific output format"""
        pass
    
    @abstractmethod
    def get_content_type(self) -> str:
        """Get MIME content type for this format"""
        pass
    
    def _serialize_datetime(self, obj):
        """Helper to serialize datetime objects"""
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        return obj

class JSONFormatter(BaseFormatter):
    """JSON output formatter"""
    
    def format(self, result: ExecutionResult) -> str:
        """Format result as JSON"""
        try:
            # Create formatted response
            formatted_data = {
                'status': result.status.value,
                'success': result.success,
                'message': result.message,
                'data': self._serialize_data(result.data),
                'metadata': {
                    'row_count': result.row_count,
                    'columns': [self._serialize_column(col) for col in result.columns],
                    'execution_timestamp': result.execution_timestamp.isoformat(),
                    'request_id': result.request_id
                }
            }
            
            # Add performance data if available
            if result.performance:
                formatted_data['performance'] = {
                    'execution_time_ms': result.performance.execution_time_ms,
                    'rows_returned': result.performance.rows_returned,
                    'connection_time_ms': result.performance.connection_time_ms
                }
            
            # Add error information if present
            if not result.success:
                formatted_data['error'] = {
                    'type': result.error_type.value if result.error_type else 'unknown',
                    'details': result.error_details,
                    'suggestions': result.suggestions
                }
            
            # Add export information
            if result.export_available:
                formatted_data['export_options'] = {
                    'available_formats': result.supported_formats,
                    'export_enabled': True
                }
            
            return json.dumps(formatted_data, indent=2, default=self._serialize_datetime, ensure_ascii=False)
            
        except Exception as e:
            # Fallback error response
            error_response = {
                'status': 'failed',
                'success': False,
                'message': f'JSON formatting error: {str(e)}',
                'data': [],
                'metadata': {'row_count': 0}
            }
            return json.dumps(error_response, indent=2)
    
    def get_content_type(self) -> str:
        return 'application/json; charset=utf-8'
    
    def _serialize_data(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Serialize data with proper datetime handling"""
        serialized = []
        for row in data:
            serialized_row = {}
            for key, value in row.items():
                serialized_row[key] = self._serialize_datetime(value)
            serialized.append(serialized_row)
        return serialized
    
    def _serialize_column(self, column: ColumnInfo) -> Dict[str, Any]:
        """Serialize column information"""
        return {
            'name': column.name,
            'data_type': column.data_type,
            'max_length': column.max_length,
            'is_nullable': column.is_nullable,
            'ordinal_position': column.ordinal_position
        }

class CSVFormatter(BaseFormatter):
    """CSV output formatter"""
    
    def format(self, result: ExecutionResult) -> str:
        """Format result as CSV"""
        try:
            if not result.data:
                return self._create_empty_csv(result)
            
            # Create CSV in memory
            output = io.StringIO()
            
            # Get column names from first row or column info
            if result.columns:
                fieldnames = [col.name for col in result.columns]
            elif result.data:
                fieldnames = list(result.data[0].keys())
            else:
                return self._create_empty_csv(result)
            
            # Create CSV writer
            writer = csv.DictWriter(
                output, 
                fieldnames=fieldnames,
                quoting=csv.QUOTE_MINIMAL,
                lineterminator='\n'
            )
            
            # Write header
            writer.writeheader()
            
            # Write data rows with proper serialization
            for row in result.data:
                serialized_row = {}
                for field in fieldnames:
                    value = row.get(field, '')
                    # Handle None values
                    if value is None:
                        serialized_row[field] = ''
                    # Handle datetime objects
                    elif isinstance(value, (datetime, date)):
                        serialized_row[field] = value.isoformat()
                    # Handle other objects
                    else:
                        serialized_row[field] = str(value)
                
                writer.writerow(serialized_row)
            
            csv_content = output.getvalue()
            output.close()
            
            return csv_content
            
        except Exception as e:
            return f"CSV formatting error: {str(e)}\n"
    
    def get_content_type(self) -> str:
        return 'text/csv; charset=utf-8'
    
    def _create_empty_csv(self, result: ExecutionResult) -> str:
        """Create CSV for empty results"""
        return f"# Query executed successfully but returned no data\n# Message: {result.message}\n# Timestamp: {result.execution_timestamp.isoformat()}\n"

class HTMLFormatter(BaseFormatter):
    """HTML table formatter"""
    
    def format(self, result: ExecutionResult) -> str:
        """Format result as HTML table"""
        try:
            if not result.data:
                return self._create_empty_html(result)
            
            # Start HTML structure
            html = ['<!DOCTYPE html>']
            html.append('<html><head>')
            html.append('<title>SQL Query Results</title>')
            html.append('<style>')
            html.append(self._get_css_styles())
            html.append('</style>')
            html.append('</head><body>')
            
            # Add header information
            html.append('<div class="header">')
            html.append(f'<h2>Query Results</h2>')
            html.append(f'<p><strong>Status:</strong> {result.status.value}</p>')
            html.append(f'<p><strong>Rows:</strong> {result.row_count:,}</p>')
            if result.performance:
                html.append(f'<p><strong>Execution Time:</strong> {result.performance.execution_time_ms}ms</p>')
            html.append(f'<p><strong>Timestamp:</strong> {result.execution_timestamp.strftime("%Y-%m-%d %H:%M:%S")}</p>')
            html.append('</div>')
            
            # Create table
            html.append('<table class="results-table">')
            
            # Table header
            if result.columns:
                html.append('<thead><tr>')
                for col in result.columns:
                    html.append(f'<th title="{col.data_type}">{self._escape_html(col.name)}</th>')
                html.append('</tr></thead>')
            
            # Table body
            html.append('<tbody>')
            for row_idx, row in enumerate(result.data):
                css_class = 'even' if row_idx % 2 == 0 else 'odd'
                html.append(f'<tr class="{css_class}">')
                
                for col in result.columns:
                    value = row.get(col.name, '')
                    formatted_value = self._format_cell_value(value)
                    html.append(f'<td>{formatted_value}</td>')
                
                html.append('</tr>')
            html.append('</tbody>')
            html.append('</table>')
            
            # Footer
            html.append('<div class="footer">')
            html.append(f'<p>Generated on {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>')
            html.append('</div>')
            
            html.append('</body></html>')
            
            return '\n'.join(html)
            
        except Exception as e:
            return f'<html><body><h2>HTML formatting error</h2><p>{self._escape_html(str(e))}</p></body></html>'
    
    def get_content_type(self) -> str:
        return 'text/html; charset=utf-8'
    
    def _create_empty_html(self, result: ExecutionResult) -> str:
        """Create HTML for empty results"""
        html = ['<!DOCTYPE html>']
        html.append('<html><head><title>No Results</title>')
        html.append('<style>body { font-family: Arial, sans-serif; margin: 40px; }</style>')
        html.append('</head><body>')
        html.append('<h2>No Data Found</h2>')
        html.append(f'<p>{result.message}</p>')
        html.append('<h3>Suggestions:</h3><ul>')
        for suggestion in result.suggestions:
            html.append(f'<li>{self._escape_html(suggestion)}</li>')
        html.append('</ul>')
        html.append('</body></html>')
        return '\n'.join(html)
    
    def _get_css_styles(self) -> str:
        """Get CSS styles for HTML table"""
        return """
            body { font-family: Arial, sans-serif; margin: 20px; }
            .header { margin-bottom: 20px; padding: 10px; background-color: #f5f5f5; border-radius: 5px; }
            .results-table { border-collapse: collapse; width: 100%; margin: 20px 0; }
            .results-table th, .results-table td { border: 1px solid #ddd; padding: 8px; text-align: left; }
            .results-table th { background-color: #4CAF50; color: white; font-weight: bold; }
            .results-table tr.even { background-color: #f9f9f9; }
            .results-table tr.odd { background-color: #ffffff; }
            .results-table tr:hover { background-color: #f5f5f5; }
            .footer { margin-top: 20px; padding: 10px; font-size: 12px; color: #666; }
        """
    
    def _escape_html(self, text: str) -> str:
        """Escape HTML special characters"""
        if text is None:
            return ''
        return str(text).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;')
    
    def _format_cell_value(self, value: Any) -> str:
        """Format cell value for HTML display"""
        if value is None:
            return '<em>NULL</em>'
        elif isinstance(value, (datetime, date)):
            return self._escape_html(value.isoformat())
        elif isinstance(value, bool):
            return 'TRUE' if value else 'FALSE'
        else:
            return self._escape_html(str(value))

class ExcelFormatter(BaseFormatter):
    """Excel output formatter using pandas and openpyxl"""
    
    def format(self, result: ExecutionResult) -> bytes:
        """Format result as Excel file"""
        try:
            if not result.data:
                return self._create_empty_excel(result)
            
            # Create DataFrame from result data
            df = pd.DataFrame(result.data)
            
            # Create Excel file in memory
            output = io.BytesIO()
            
            # Create Excel writer with openpyxl engine
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Write main data
                df.to_excel(writer, sheet_name='Query Results', index=False)
                
                # Add metadata sheet
                metadata_data = {
                    'Property': [
                        'Query Status',
                        'Row Count',
                        'Execution Time (ms)',
                        'Execution Timestamp',
                        'Request ID'
                    ],
                    'Value': [
                        result.status.value,
                        result.row_count,
                        result.performance.execution_time_ms if result.performance else 'N/A',
                        result.execution_timestamp.isoformat(),
                        result.request_id or 'N/A'
                    ]
                }
                
                metadata_df = pd.DataFrame(metadata_data)
                metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
                
                # Format the worksheets
                self._format_excel_worksheets(writer)
            
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            # Create error Excel file
            error_data = {'Error': [f'Excel formatting error: {str(e)}']}
            error_df = pd.DataFrame(error_data)
            output = io.BytesIO()
            error_df.to_excel(output, index=False, engine='openpyxl')
            output.seek(0)
            return output.getvalue()
    
    def get_content_type(self) -> str:
        return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    def _create_empty_excel(self, result: ExecutionResult) -> bytes:
        """Create Excel file for empty results"""
        data = {
            'Message': [result.message],
            'Suggestions': result.suggestions
        }
        
        # Pad suggestions list to match message length
        max_len = max(len(data['Message']), len(data['Suggestions']))
        for key in data:
            while len(data[key]) < max_len:
                data[key].append('')
        
        df = pd.DataFrame(data)
        output = io.BytesIO()
        df.to_excel(output, index=False, engine='openpyxl')
        output.seek(0)
        return output.getvalue()
    
    def _format_excel_worksheets(self, writer):
        """Apply formatting to Excel worksheets"""
        try:
            from openpyxl.styles import Font, PatternFill
            
            # Format the main results sheet
            workbook = writer.book
            if 'Query Results' in workbook.sheetnames:
                worksheet = workbook['Query Results']
                
                # Format header row
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                
                for cell in worksheet[1]:  # First row
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        except ImportError:
            # openpyxl styles not available, skip formatting
            pass
        except Exception:
            # Formatting failed, continue without formatting
            pass

class ResultFormatter:
    """Main result formatter that delegates to specific formatters"""
    
    def __init__(self):
        self.formatters = {
            OutputFormat.JSON: JSONFormatter(),
            OutputFormat.CSV: CSVFormatter(),
            OutputFormat.HTML: HTMLFormatter(),
            OutputFormat.EXCEL: ExcelFormatter()
        }
    
    def format_result(self, result: ExecutionResult, output_format: OutputFormat) -> Union[str, bytes]:
        """Format result in specified format"""
        try:
            formatter = self.formatters.get(output_format)
            if not formatter:
                raise ValueError(f"Unsupported output format: {output_format}")
            
            return formatter.format(result)
            
        except Exception as e:
            # Fallback to JSON error response
            error_result = {
                'status': 'failed',
                'success': False,
                'message': f'Formatting error: {str(e)}',
                'data': [],
                'error_type': 'formatting_error'
            }
            return json.dumps(error_result, indent=2)
    
    def get_content_type(self, output_format: OutputFormat) -> str:
        """Get content type for format"""
        formatter = self.formatters.get(output_format)
        return formatter.get_content_type() if formatter else 'application/json'
    
    def get_supported_formats(self) -> List[str]:
        """Get list of supported format names"""
        return [fmt.value for fmt in self.formatters.keys()]
    
    def format_multiple(self, result: ExecutionResult, formats: List[OutputFormat]) -> Dict[OutputFormat, Union[str, bytes]]:
        """Format result in multiple formats"""
        formatted_results = {}
        
        for fmt in formats:
            try:
                formatted_results[fmt] = self.format_result(result, fmt)
            except Exception as e:
                # Store error for this format
                formatted_results[fmt] = f"Error formatting {fmt.value}: {str(e)}"
        
        return formatted_results

# Convenience functions
def format_as_json(result: ExecutionResult) -> str:
    """Quick JSON formatting"""
    formatter = JSONFormatter()
    return formatter.format(result)

def format_as_csv(result: ExecutionResult) -> str:
    """Quick CSV formatting"""
    formatter = CSVFormatter()
    return formatter.format(result)

def format_as_html(result: ExecutionResult) -> str:
    """Quick HTML formatting"""
    formatter = HTMLFormatter()
    return formatter.format(result)

def format_as_excel(result: ExecutionResult) -> bytes:
    """Quick Excel formatting"""
    formatter = ExcelFormatter()
    return formatter.format(result)

# Export main classes and functions
__all__ = [
    'BaseFormatter', 'JSONFormatter', 'CSVFormatter', 'HTMLFormatter', 'ExcelFormatter',
    'ResultFormatter',
    'format_as_json', 'format_as_csv', 'format_as_html', 'format_as_excel'
]
